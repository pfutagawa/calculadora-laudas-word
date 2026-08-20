"""Calculadora desktop de laudas e orçamentos para documentos Word.

Requisitos no Windows: ``pip install pywin32 openpyxl``.
"""

from __future__ import annotations

import os
import queue
import threading
from datetime import datetime
from pathlib import Path
from typing import Any
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import pythoncom
    import win32com.client as win32
except ImportError:
    pythoncom = None
    win32 = None


APP_TITLE = "Calculadora de Laudas e Orçamento"
WORD_EXTENSIONS = {".doc", ".docx"}


def formatar_numero(valor: float, casas: int = 2) -> str:
    """Formata números no padrão brasileiro sem depender do locale do Windows."""
    texto = f"{valor:,.{casas}f}"
    return texto.replace(",", "§").replace(".", ",").replace("§", ".")


def formatar_moeda(valor: float) -> str:
    return f"R$ {formatar_numero(valor)}"


def converter_decimal(texto: str) -> float:
    """Aceita 35,50, 35.50 e números com separador de milhar."""
    valor = texto.strip().replace("R$", "").replace(" ", "")
    if not valor:
        raise ValueError

    if "," in valor:
        valor = valor.replace(".", "").replace(",", ".")
    return float(valor)


class CalculadoraLaudas(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1120x720")
        self.minsize(900, 600)
        self.configure(bg="#f4f6f8")

        self.pasta_var = tk.StringVar()
        self.valor_lauda_var = tk.StringVar(value="35,00")
        self.caracteres_lauda_var = tk.StringVar(value="1000")
        self.incluir_subpastas_var = tk.BooleanVar(value=True)
        self.status_var = tk.StringVar(value="Selecione uma pasta para começar.")

        self.fila: queue.Queue[tuple[str, Any]] = queue.Queue()
        self.cancelar_evento = threading.Event()
        self.processando = False
        self.resultados: list[dict[str, Any]] = []
        self.valor_lauda_atual = 0.0
        self.caracteres_lauda_atual = 1000

        self._configurar_estilos()
        self._montar_interface()
        self.protocol("WM_DELETE_WINDOW", self._ao_fechar)

    def _configurar_estilos(self) -> None:
        estilo = ttk.Style(self)
        if "vista" in estilo.theme_names():
            estilo.theme_use("vista")
        elif "clam" in estilo.theme_names():
            estilo.theme_use("clam")

        estilo.configure("App.TFrame", background="#f4f6f8")
        estilo.configure("Card.TFrame", background="#ffffff")
        estilo.configure(
            "Titulo.TLabel",
            background="#f4f6f8",
            foreground="#17202a",
            font=("Segoe UI", 19, "bold"),
        )
        estilo.configure(
            "Subtitulo.TLabel",
            background="#f4f6f8",
            foreground="#5f6b76",
            font=("Segoe UI", 10),
        )
        estilo.configure(
            "CardTitulo.TLabel",
            background="#ffffff",
            foreground="#5f6b76",
            font=("Segoe UI", 9),
        )
        estilo.configure(
            "CardValor.TLabel",
            background="#ffffff",
            foreground="#17202a",
            font=("Segoe UI", 15, "bold"),
        )
        estilo.configure("Treeview", rowheight=27, font=("Segoe UI", 9))
        estilo.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        estilo.configure("Accent.TButton", font=("Segoe UI", 10, "bold"))

    def _montar_interface(self) -> None:
        principal = ttk.Frame(self, style="App.TFrame", padding=24)
        principal.pack(fill="both", expand=True)
        principal.columnconfigure(0, weight=1)
        principal.rowconfigure(4, weight=1)

        ttk.Label(principal, text=APP_TITLE, style="Titulo.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(
            principal,
            text="Conte os caracteres de documentos Word e gere o orçamento em Excel.",
            style="Subtitulo.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(2, 18))

        configuracao = ttk.LabelFrame(
            principal, text=" Configuração ", padding=(16, 12)
        )
        configuracao.grid(row=2, column=0, sticky="ew")
        configuracao.columnconfigure(0, weight=1)

        ttk.Label(configuracao, text="Pasta com os documentos Word").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 5)
        )
        self.entrada_pasta = ttk.Entry(
            configuracao, textvariable=self.pasta_var, state="readonly"
        )
        self.entrada_pasta.grid(row=1, column=0, sticky="ew", padx=(0, 8))
        self.botao_pasta = ttk.Button(
            configuracao, text="Selecionar pasta…", command=self._selecionar_pasta
        )
        self.botao_pasta.grid(row=1, column=1, sticky="ew")

        ttk.Label(configuracao, text="Valor por lauda (R$)").grid(
            row=2, column=0, sticky="w", pady=(12, 5)
        )
        ttk.Label(configuracao, text="Caracteres por lauda").grid(
            row=2, column=1, sticky="w", padx=(12, 0), pady=(12, 5)
        )

        self.entrada_valor = ttk.Entry(
            configuracao, textvariable=self.valor_lauda_var, width=18
        )
        self.entrada_valor.grid(row=3, column=0, sticky="w")
        self.entrada_caracteres = ttk.Entry(
            configuracao, textvariable=self.caracteres_lauda_var, width=18
        )
        self.entrada_caracteres.grid(row=3, column=1, sticky="w", padx=(12, 0))
        self.check_subpastas = ttk.Checkbutton(
            configuracao,
            text="Incluir documentos das subpastas",
            variable=self.incluir_subpastas_var,
        )
        self.check_subpastas.grid(row=3, column=2, sticky="w", padx=(16, 0))

        botoes = ttk.Frame(configuracao)
        botoes.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(16, 0))
        self.botao_processar = ttk.Button(
            botoes,
            text="Processar documentos",
            command=self._iniciar_processamento,
            style="Accent.TButton",
        )
        self.botao_processar.pack(side="left")
        self.botao_cancelar = ttk.Button(
            botoes,
            text="Cancelar",
            command=self._cancelar_processamento,
            state="disabled",
        )
        self.botao_cancelar.pack(side="left", padx=(8, 0))
        self.botao_limpar = ttk.Button(
            botoes, text="Limpar resultados", command=self._limpar_resultados
        )
        self.botao_limpar.pack(side="right")

        resumo = ttk.Frame(principal, style="App.TFrame")
        resumo.grid(row=3, column=0, sticky="ew", pady=16)
        for coluna in range(4):
            resumo.columnconfigure(coluna, weight=1, uniform="cards")

        self.card_documentos = self._criar_card(resumo, 0, "DOCUMENTOS", "0")
        self.card_caracteres = self._criar_card(resumo, 1, "CARACTERES", "0")
        self.card_laudas = self._criar_card(resumo, 2, "LAUDAS", "0,00")
        self.card_total = self._criar_card(resumo, 3, "TOTAL", "R$ 0,00")

        area_tabela = ttk.Frame(principal)
        area_tabela.grid(row=4, column=0, sticky="nsew")
        area_tabela.columnconfigure(0, weight=1)
        area_tabela.rowconfigure(0, weight=1)

        colunas = ("arquivo", "pasta", "caracteres", "laudas", "valor", "status")
        self.tabela = ttk.Treeview(
            area_tabela, columns=colunas, show="headings", selectmode="browse"
        )
        titulos = {
            "arquivo": "Arquivo",
            "pasta": "Pasta",
            "caracteres": "Caracteres",
            "laudas": "Laudas",
            "valor": "Valor",
            "status": "Status",
        }
        larguras = {
            "arquivo": 220,
            "pasta": 290,
            "caracteres": 100,
            "laudas": 85,
            "valor": 100,
            "status": 150,
        }
        for coluna in colunas:
            self.tabela.heading(coluna, text=titulos[coluna])
            ancora = "e" if coluna in {"caracteres", "laudas", "valor"} else "w"
            self.tabela.column(
                coluna,
                width=larguras[coluna],
                minwidth=70,
                anchor=ancora,
                stretch=coluna in {"arquivo", "pasta", "status"},
            )

        self.tabela.tag_configure("erro", foreground="#b42318")
        self.tabela.grid(row=0, column=0, sticky="nsew")
        rolagem_y = ttk.Scrollbar(
            area_tabela, orient="vertical", command=self.tabela.yview
        )
        rolagem_y.grid(row=0, column=1, sticky="ns")
        rolagem_x = ttk.Scrollbar(
            area_tabela, orient="horizontal", command=self.tabela.xview
        )
        rolagem_x.grid(row=1, column=0, sticky="ew")
        self.tabela.configure(
            yscrollcommand=rolagem_y.set, xscrollcommand=rolagem_x.set
        )

        rodape = ttk.Frame(principal, style="App.TFrame")
        rodape.grid(row=5, column=0, sticky="ew", pady=(12, 0))
        rodape.columnconfigure(0, weight=1)
        ttk.Label(
            rodape, textvariable=self.status_var, style="Subtitulo.TLabel"
        ).grid(row=0, column=0, sticky="w")
        self.botao_salvar = ttk.Button(
            rodape,
            text="Salvar planilha Excel…",
            command=self._salvar_excel,
            state="disabled",
        )
        self.botao_salvar.grid(row=0, column=1, sticky="e", padx=(12, 0))
        self.progresso = ttk.Progressbar(rodape, mode="determinate")
        self.progresso.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(8, 0))

    def _criar_card(
        self, pai: ttk.Frame, coluna: int, titulo: str, valor: str
    ) -> ttk.Label:
        card = ttk.Frame(pai, style="Card.TFrame", padding=(14, 10))
        card.grid(
            row=0,
            column=coluna,
            sticky="ew",
            padx=(0 if coluna == 0 else 6, 0 if coluna == 3 else 6),
        )
        ttk.Label(card, text=titulo, style="CardTitulo.TLabel").pack(anchor="w")
        label_valor = ttk.Label(card, text=valor, style="CardValor.TLabel")
        label_valor.pack(anchor="w", pady=(3, 0))
        return label_valor

    def _selecionar_pasta(self) -> None:
        pasta = filedialog.askdirectory(
            parent=self,
            title="Selecione a pasta com os documentos Word",
            initialdir=self.pasta_var.get() or str(Path.home()),
        )
        if pasta:
            self.pasta_var.set(os.path.normpath(pasta))
            self.status_var.set("Pasta selecionada. Clique em ‘Processar documentos’.")

    def _validar_parametros(self) -> tuple[Path, float, int] | None:
        pasta = Path(self.pasta_var.get())
        if not self.pasta_var.get() or not pasta.is_dir():
            messagebox.showwarning(APP_TITLE, "Selecione uma pasta válida.", parent=self)
            return None

        try:
            valor_lauda = converter_decimal(self.valor_lauda_var.get())
            if valor_lauda < 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning(
                APP_TITLE,
                "Informe um valor por lauda válido (por exemplo, 35,00).",
                parent=self,
            )
            self.entrada_valor.focus_set()
            return None

        try:
            caracteres_lauda = int(self.caracteres_lauda_var.get().strip())
            if caracteres_lauda <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning(
                APP_TITLE,
                "Informe uma quantidade inteira e positiva de caracteres por lauda.",
                parent=self,
            )
            self.entrada_caracteres.focus_set()
            return None

        return pasta, valor_lauda, caracteres_lauda

    def _listar_documentos(self, pasta: Path) -> list[Path]:
        padrao = pasta.rglob("*") if self.incluir_subpastas_var.get() else pasta.glob("*")
        return sorted(
            (
                caminho.resolve()
                for caminho in padrao
                if caminho.is_file()
                and caminho.suffix.lower() in WORD_EXTENSIONS
                and not caminho.name.startswith("~$")
            ),
            key=lambda caminho: str(caminho).casefold(),
        )

    def _iniciar_processamento(self) -> None:
        parametros = self._validar_parametros()
        if not parametros:
            return

        pasta, valor_lauda, caracteres_lauda = parametros
        documentos = self._listar_documentos(pasta)
        if not documentos:
            messagebox.showinfo(
                APP_TITLE,
                "Nenhum arquivo .doc ou .docx foi encontrado na pasta selecionada.",
                parent=self,
            )
            return

        self._limpar_resultados(confirmar=False)
        self.valor_lauda_atual = valor_lauda
        self.caracteres_lauda_atual = caracteres_lauda
        self.cancelar_evento.clear()
        self.processando = True
        self.progresso.configure(maximum=len(documentos), value=0)
        self.status_var.set(f"Preparando o Microsoft Word… 0 de {len(documentos)}")
        self._alternar_controles(processando=True)

        thread = threading.Thread(
            target=self._processar_documentos,
            args=(documentos, valor_lauda, caracteres_lauda),
            daemon=True,
        )
        thread.start()
        self.after(100, self._ler_fila)

    def _processar_documentos(
        self, documentos: list[Path], valor_lauda: float, caracteres_lauda: int
    ) -> None:
        if pythoncom is None or win32 is None:
            self.fila.put(
                (
                    "fatal",
                    "O componente pywin32 não está instalado. Abra o Prompt de Comando "
                    "e execute: pip install pywin32 openpyxl",
                )
            )
            return

        word = None
        pythoncom.CoInitialize()
        try:
            try:
                word = win32.DispatchEx("Word.Application")
                word.Visible = False
                word.DisplayAlerts = 0
            except Exception as erro:
                self.fila.put(("fatal", f"Não foi possível iniciar o Microsoft Word: {erro}"))
                return

            for indice, caminho in enumerate(documentos, start=1):
                if self.cancelar_evento.is_set():
                    self.fila.put(("cancelado", None))
                    return

                doc = None
                resultado: dict[str, Any] = {
                    "arquivo": caminho.name,
                    "pasta": str(caminho.parent),
                    "caminho": str(caminho),
                    "caracteres": None,
                    "laudas": None,
                    "valor": None,
                    "status": "",
                }
                try:
                    doc = word.Documents.Open(
                        str(caminho),
                        ConfirmConversions=False,
                        ReadOnly=True,
                        AddToRecentFiles=False,
                    )
                    # 5 = wdStatisticCharactersWithSpaces
                    caracteres = int(doc.ComputeStatistics(5, True))
                    laudas = caracteres / caracteres_lauda
                    resultado.update(
                        {
                            "caracteres": caracteres,
                            "laudas": laudas,
                            "valor": laudas * valor_lauda,
                            "status": "Processado",
                        }
                    )
                except Exception as erro:
                    mensagem = str(erro).strip().replace("\n", " ")
                    resultado["status"] = f"Erro: {mensagem or 'falha ao abrir o arquivo'}"
                finally:
                    if doc is not None:
                        try:
                            doc.Close(False)
                        except Exception:
                            pass

                self.fila.put(("resultado", (resultado, indice, len(documentos))))

            self.fila.put(("concluido", None))
        finally:
            if word is not None:
                try:
                    word.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()

    def _ler_fila(self) -> None:
        try:
            while True:
                evento, conteudo = self.fila.get_nowait()
                if evento == "resultado":
                    resultado, indice, total = conteudo
                    self.resultados.append(resultado)
                    self._adicionar_resultado_tabela(resultado)
                    self.progresso.configure(value=indice)
                    self.status_var.set(f"Processando {resultado['arquivo']} — {indice} de {total}")
                    self._atualizar_resumo()
                elif evento == "concluido":
                    self._finalizar_processamento(cancelado=False)
                elif evento == "cancelado":
                    self._finalizar_processamento(cancelado=True)
                elif evento == "fatal":
                    self._finalizar_processamento(cancelado=False, erro_fatal=str(conteudo))
        except queue.Empty:
            pass

        if self.processando:
            self.after(100, self._ler_fila)

    def _adicionar_resultado_tabela(self, resultado: dict[str, Any]) -> None:
        sucesso = resultado["caracteres"] is not None
        valores = (
            resultado["arquivo"],
            resultado["pasta"],
            formatar_numero(resultado["caracteres"], 0) if sucesso else "—",
            formatar_numero(resultado["laudas"]) if sucesso else "—",
            formatar_moeda(resultado["valor"]) if sucesso else "—",
            resultado["status"],
        )
        self.tabela.insert("", "end", values=valores, tags=(() if sucesso else ("erro",)))

    def _atualizar_resumo(self) -> None:
        validos = [r for r in self.resultados if r["caracteres"] is not None]
        total_caracteres = sum(r["caracteres"] for r in validos)
        total_laudas = sum(r["laudas"] for r in validos)
        total_valor = sum(r["valor"] for r in validos)

        self.card_documentos.configure(text=str(len(validos)))
        self.card_caracteres.configure(text=formatar_numero(total_caracteres, 0))
        self.card_laudas.configure(text=formatar_numero(total_laudas))
        self.card_total.configure(text=formatar_moeda(total_valor))

    def _finalizar_processamento(
        self, cancelado: bool, erro_fatal: str | None = None
    ) -> None:
        self.processando = False
        self._alternar_controles(processando=False)
        self.botao_salvar.configure(state="normal" if self.resultados else "disabled")

        if erro_fatal:
            self.status_var.set("Não foi possível iniciar o processamento.")
            messagebox.showerror(APP_TITLE, erro_fatal, parent=self)
            return

        validos = sum(r["caracteres"] is not None for r in self.resultados)
        erros = len(self.resultados) - validos
        if cancelado:
            self.status_var.set(
                f"Processamento cancelado. {validos} documento(s) concluído(s)."
            )
        elif erros:
            self.status_var.set(
                f"Concluído: {validos} documento(s) processado(s) e {erros} com erro."
            )
        else:
            self.status_var.set(f"Concluído: {validos} documento(s) processado(s).")

    def _alternar_controles(self, processando: bool) -> None:
        estado_normal = "disabled" if processando else "normal"
        self.botao_pasta.configure(state=estado_normal)
        self.botao_processar.configure(state=estado_normal)
        self.botao_limpar.configure(state=estado_normal)
        self.entrada_valor.configure(state=estado_normal)
        self.entrada_caracteres.configure(state=estado_normal)
        self.check_subpastas.configure(state=estado_normal)
        self.botao_cancelar.configure(state="normal" if processando else "disabled")
        if processando:
            self.botao_salvar.configure(state="disabled")

    def _cancelar_processamento(self) -> None:
        self.cancelar_evento.set()
        self.botao_cancelar.configure(state="disabled")
        self.status_var.set("Cancelamento solicitado; concluindo o arquivo atual…")

    def _limpar_resultados(self, confirmar: bool = True) -> None:
        if confirmar and self.resultados:
            if not messagebox.askyesno(
                APP_TITLE, "Deseja limpar os resultados atuais?", parent=self
            ):
                return
        self.resultados.clear()
        for item in self.tabela.get_children():
            self.tabela.delete(item)
        self.progresso.configure(value=0)
        self.botao_salvar.configure(state="disabled")
        self._atualizar_resumo()
        if confirmar:
            self.status_var.set("Resultados limpos.")

    def _salvar_excel(self) -> None:
        if not self.resultados:
            return

        nome_padrao = f"Orcamento_Sesc_{datetime.now():%Y-%m-%d}.xlsx"
        destino = filedialog.asksaveasfilename(
            parent=self,
            title="Salvar planilha de orçamento",
            initialdir=self.pasta_var.get() or str(Path.home()),
            initialfile=nome_padrao,
            defaultextension=".xlsx",
            filetypes=[("Planilha do Excel", "*.xlsx")],
        )
        if not destino:
            return

        try:
            self._gerar_excel(Path(destino))
        except Exception as erro:
            messagebox.showerror(
                APP_TITLE, f"Não foi possível salvar a planilha:\n\n{erro}", parent=self
            )
            return

        self.status_var.set(f"Planilha salva em: {destino}")
        messagebox.showinfo(APP_TITLE, "Planilha salva com sucesso.", parent=self)

    def _gerar_excel(self, destino: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Orçamento"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:G1"

        cabecalhos = [
            "Caminho da Pasta",
            "Nome do Arquivo",
            "Caracteres (com espaços)",
            "Caracteres por Lauda",
            "Total de Laudas",
            "Valor",
            "Status",
        ]
        ws.append(cabecalhos)

        for resultado in self.resultados:
            ws.append(
                [
                    resultado["pasta"],
                    resultado["arquivo"],
                    resultado["caracteres"],
                    self.caracteres_lauda_atual,
                    resultado["laudas"],
                    resultado["valor"],
                    resultado["status"],
                ]
            )

        validos = [r for r in self.resultados if r["caracteres"] is not None]
        total_caracteres = sum(r["caracteres"] for r in validos)
        total_laudas = sum(r["laudas"] for r in validos)
        total_valor = sum(r["valor"] for r in validos)

        ws.append([])
        linha_resumo = ws.max_row + 1
        resumos = [
            ("TOTAL DE DOCUMENTOS", len(validos)),
            ("TOTAL DE CARACTERES", total_caracteres),
            (f"TOTAL DE LAUDAS ({self.caracteres_lauda_atual})", total_laudas),
            ("VALOR POR LAUDA", self.valor_lauda_atual),
            ("TOTAL DO ORÇAMENTO", total_valor),
        ]
        for rotulo, valor in resumos:
            ws.append(["", rotulo, valor])

        preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
        for celula in ws[1]:
            celula.fill = preenchimento_cabecalho
            celula.font = Font(color="FFFFFF", bold=True)
            celula.alignment = Alignment(horizontal="center", vertical="center")

        for linha in range(2, 2 + len(self.resultados)):
            ws.cell(linha, 3).number_format = "#,##0"
            ws.cell(linha, 4).number_format = "#,##0"
            ws.cell(linha, 5).number_format = "#,##0.00"
            ws.cell(linha, 6).number_format = 'R$ #,##0.00'

        for linha in range(linha_resumo, linha_resumo + len(resumos)):
            ws.cell(linha, 2).font = Font(bold=True)
            ws.cell(linha, 2).fill = PatternFill("solid", fgColor="D9EAF7")
            ws.cell(linha, 3).font = Font(bold=True)
        ws.cell(linha_resumo + 2, 3).number_format = "#,##0.00"
        ws.cell(linha_resumo + 3, 3).number_format = 'R$ #,##0.00'
        ws.cell(linha_resumo + 4, 3).number_format = 'R$ #,##0.00'

        larguras = [42, 34, 24, 22, 18, 16, 42]
        for indice, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(indice)].width = largura

        ws.row_dimensions[1].height = 28
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)

    def _ao_fechar(self) -> None:
        if self.processando:
            fechar = messagebox.askyesno(
                APP_TITLE,
                "Há um processamento em andamento. Deseja cancelá-lo e fechar?",
                parent=self,
            )
            if not fechar:
                return
            self.cancelar_evento.set()
        self.destroy()


def configurar_dpi_windows() -> None:
    """Melhora a nitidez da interface em telas Windows com escala elevada."""
    try:
        from ctypes import windll

        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass


if __name__ == "__main__":
    configurar_dpi_windows()
    app = CalculadoraLaudas()
    app.mainloop()
